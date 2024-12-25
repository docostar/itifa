from openpyxl import load_workbook, Workbook
from fun import get_practical_number,get_practical_wise_marks, get_section_marks, increment_column, distribute_marks
import shutil

iti_name = "BAGASARA(MAHILA)"

# Give the location of the file
input_file = "FA_Input.xlsx"
demo_file= "FA_Output.xlsx"

# To open the workbook
# workbook object is created
wb_input = load_workbook(input_file)

sheet = wb_input["Instructor"]
si_name = sheet["B1"].value
year = sheet["B2"].value
sem = sheet["B3"].value
batch = sheet["B4"].value
trade = sheet["B5"].value
duration = sheet["B6"].value
location = "BAGASARA"

sheet_name = wb_input["Name_Rollno"]
student_start_row = 2
student_end_row = sheet_name.max_row
sheet_lo = wb_input["LO"]
lo_no = sheet_name["G2"].value

output_file="LO_"+str(lo_no)+".xlsx"
shutil.copy(demo_file, output_file)
print(f"File {output_file} created successfully.")
wb_out = load_workbook(output_file)
marks_start_row = 9
start_practical, end_practical = get_practical_number(sheet_lo,lo_no)
no_of_practical = end_practical - start_practical + 1


for student_row_no in range(student_start_row,student_end_row+1):
    student_row_str=str(student_row_no)
    student_name = sheet_name["B"+student_row_str].value
    Roll_No = sheet_name["A"+student_row_str].value
    student_main_mark = sheet_name["C"+student_row_str].value

    sheet_to_copy = wb_out["demo"]
    student_sheet = wb_out.copy_worksheet(sheet_to_copy)
    student_sheet.title = str(Roll_No)

    student_sheet["E2"].value = student_name
    student_sheet["U2"].value = Roll_No
    student_sheet["AD2"].value = year
    student_sheet["AL2"].value = sem
    student_sheet["E3"].value = iti_name
    student_sheet["AL3"].value = batch
    student_sheet["AD4"].value = location
    student_sheet["E5"].value = trade
    student_sheet["AD5"].value = duration
    student_sheet["AK5"].value = si_name

    practical_no = start_practical
    practical_marks = get_practical_wise_marks(student_main_mark,no_of_practical)
    for row_no in range(marks_start_row,marks_start_row+no_of_practical):
        row_no_str = str(row_no)
        student_sheet["A"+row_no_str] = "LO"+str(lo_no)
        student_sheet["B"+row_no_str] = practical_no
        total_practical_marks =  practical_marks[practical_no - start_practical]
        student_sheet["AM"+row_no_str] = total_practical_marks

        section_marks = distribute_marks(total_marks=total_practical_marks )
        col = "C"
        for smark in section_marks:
            for _ in smark:
                student_sheet[col+row_no_str] = _
                col = increment_column(col)
            student_sheet[col+row_no_str] = sum(smark)
            col = increment_column(col)
        col = increment_column(col)
        col = increment_column(col)


        practical_no = practical_no + 1
        

wb_out.save(output_file)